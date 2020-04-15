# -*- coding:utf-8 -*-
"""
==============================
Author:yuan
Time  :2020/3/18 22:18
file  :handler_doexcel.py
===============================
"""
import openpyxl



class DoExcle():

    def __init__(self, filapath, sheet_name):
        '''初始化用例文件及表单'''

        self.filepath = filapath
        self.sheet_name = sheet_name

    def read_excle_data_is_dict_01(self):
        '''
        读取测试数据,传入的数据是字典
        '''
        # 创建工作薄对象
        wb = openpyxl.load_workbook(self.filepath)
        # 生成需要操作的表单对象
        sheet = wb[self.sheet_name]
        test_data = []
        # 按行读取数据i行1,2,3,4,5列的数据
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data['case_id'] = sheet.cell(i, 1).value
            sub_data['title'] = sheet.cell(i, 3).value
            sub_data['method'] = sheet.cell(i, 4).value
            sub_data['url'] = sheet.cell(i, 5).value
            sub_data['data'] = sheet.cell(i, 6).value
            sub_data['expected'] = sheet.cell(i, 7).value
            sub_data['check_sql'] = sheet.cell(i,10).value
            test_data.append(sub_data)
        return test_data

    def read_excle_data_is_dict_02(self):
        '''
        读取测试数据，传入的数据是字典
        :return:
        '''
        # 获取操作文档对象
        wb = openpyxl.load_workbook(self.filepath)
        # 获取sheet表单操作对象
        sheet = wb[self.sheet_name]
        # 按行读取文档内容
        sheet_data = sheet.rows
        # 准备接收标题数据
        title_list = []
        # 遍历整个sheet表单获取每行对象
        for item in sheet_data:
            row_list = []
            # 遍历每行的数据，获取单元格对象的数据
            for i in item:
                row_list.append(i.value)
            # 每行数据生成一个列表
            title_list.append(row_list)
        # 接收生成字典需要的值
        data_dict_value = []
        # 指定标题为字典的键
        data_dict_key = title_list[0]
        # 遍历出列表对应的数据值，每个单元格数据生成一个猎列表，并与键打包成一个对象
        for i in title_list[1:]:
            aa = zip(data_dict_key, i)
            data_dict_value.append(dict(aa))
        return data_dict_value

    def read_excle_data_is_list(self):
        '''
        读取测试数据，读取的是列表形式
        :return:
        '''
        wb = openpyxl.load_workbook(self.filepath)
        sheet = wb[self.sheet_name]
        # 把按行读取的对象转换为列表
        datas = list(sheet.rows)
        # 创建空列表，接收测试用例整体数据
        case_datas = []
        # 创建空列表，接收测试用例表头值
        title = []
        # 遍历datas第一行的数据加入列表，得到title
        for item in datas[0]:
            title.append(item.value)
        # 读取用例的数据信息
        for case in datas[1:]:
            # 创建空列表，接收用例数据
            case_data_list = []
            for case_data in case:
                case_data_list.append(case_data.value)
                # 将title与case_data_list打包成字典
                case_dict = dict(zip(title, case_data_list))
            case_datas.append(case_dict)
        return case_datas

    def back_write(self, row_number, column_number, value):
        '''
        把测试结果写入测试用例文档
        :param row_number: 行
        :param column_number:列
        :param value:数据值
        :return:
        '''
        wb = openpyxl.load_workbook(self.filepath)
        sheet = wb[self.sheet_name]
        # 在指定行写入值
        sheet.cell(row_number, column_number).value = value
        # 写完以后需要保存工作薄对象为文件
        wb.save(self.filepath)


if __name__ == '__main__':
    # print(DoExcle('case_data.xlsx', 'login').read_excle())
    # print(DoExcle('case_data.xlsx', 'login').read_casedata())
    print(DoExcle('cases.xlsx', 'register').read_excle_data_is_list())
